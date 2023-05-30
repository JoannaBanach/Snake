import xlsxwriter
import openpyxl
import os
from datetime import date


class SaveScore:
    def __init__(self):
        self.name = "GameSnakeScore.xlsx"
        self.levels = ['Super Easy', 'Easy', 'Medium', 'Hard', 'Super Hard']

    def initiate_file(self):
        self.create_file()
        self.check_workbooks()

    def check_exist(self):
        result = os.path.exists(self.name)
        return result

    def create_file(self):
        if not self.check_exist():
            workbook = xlsxwriter.Workbook(self.name)
            workbook.close()

    def check_workbooks(self):
        workbook = openpyxl.load_workbook(self.name)
        workbook.active
        for sheet in self.levels:
            if sheet not in list(workbook.sheetnames):
                workbook.create_sheet(sheet)
                workbook[sheet].cell(row=1, column=1, value='Place')
                workbook[sheet].cell(row=1, column=2, value='Date')
                workbook[sheet].cell(row=1, column=3, value='Player')
                workbook[sheet].cell(row=1, column=4, value='Score')
        workbook.save(self.name)
        workbook.close()

    def write_score(self, score, game_level, player_name):
        workbook = openpyxl.load_workbook(self.name)
        workbook.active
        sheet = game_level
        current_result = workbook[sheet]["A2:D11"]
        current_result = current_result[::-1]

        current_last_place = 0
        current_place = 11

        for c1, c2, c3, c4 in current_result:
            if not c1.value:
                continue
            elif c1.value in range(1, 11) and current_last_place == 0:
                current_last_place = c1.value

            if current_last_place != 0:
                if c4.value and c4.value < score:
                    current_place = c1.value

        if current_place == 11:
            current_place = current_last_place + 1

        if current_place == current_last_place + 1 and current_place < 11:
            workbook[sheet].cell(row=current_place + 1, column=1, value=current_place)
            workbook[sheet].cell(row=current_place + 1, column=2, value=date.today().strftime('%Y-%m-%d'))
            workbook[sheet].cell(row=current_place + 1, column=3, value=player_name)
            workbook[sheet].cell(row=current_place+1, column=4, value=score)

        elif current_place < 11:
            for i in range(current_last_place+1, current_place, -1):
                workbook[sheet].cell(row=i + 1, column=1, value=i)
                try:
                    workbook[sheet].cell(row=i + 1, column=2,
                                         value=workbook[sheet].cell(row=i, column=2).value.strftime('%Y-%m-%d'))
                except AttributeError:
                    workbook[sheet].cell(row=i + 1, column=2,
                                         value=workbook[sheet].cell(row=i, column=2).value)
                workbook[sheet].cell(row=i + 1, column=3, value=workbook[sheet].cell(row=i, column=3).value)
                workbook[sheet].cell(row=i + 1, column=4, value=workbook[sheet].cell(row=i, column=4).value)
            workbook[sheet].cell(row=current_place + 1, column=1, value=current_place)
            workbook[sheet].cell(row=current_place + 1, column=2, value=date.today().strftime('%Y-%m-%d'))
            workbook[sheet].cell(row=current_place + 1, column=3, value=player_name)
            workbook[sheet].cell(row=current_place + 1, column=4, value=score)

        workbook.save(self.name)
        workbook.close()

        return current_place
